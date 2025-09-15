import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

# Raw text data
raw_text = """
Prax's Restaurant
Address: G floor, I-Rise Building, First Al Khail Road - Hessa Street,Tecom, Dubai, UAE
Modern dining with a creative twist.
  Verified+10  Years with us  Updated
+971 800 77297E-mailMapWebsite4 Photos
...
"""  # Add all your data here exactly as you have it.

# Step 1: Split into lines and remove empty lines
lines = raw_text.strip().split('\n')
clean_lines = [line.strip() for line in lines if line.strip() != ""]

# Step 2: Remove lines containing 'Review'
filtered_lines = [line for line in clean_lines if "Review" not in line]

# Step 3: Group lines where each group ends at a line containing 'Photos'
grouped_data = []
temp_group = []

for line in filtered_lines:
    temp_group.append(line)
    if "Photos" in line:
        grouped_data.append(temp_group)
        temp_group = []

if temp_group:
    grouped_data.append(temp_group)

# Step 4: Find maximum columns
max_cols = max(len(group) for group in grouped_data)

# Step 5: Pad groups to have equal length
for group in grouped_data:
    while len(group) < max_cols:
        group.append("")

# Step 6: Create DataFrame
column_names = [f"Field {i+1}" for i in range(max_cols)]
df = pd.DataFrame(grouped_data, columns=column_names)

# Step 7: Remove 'Field 4'
if "Field 4" in df.columns:
    df = df.drop(columns=["Field 4"])

# Step 8: Process 'Field 5'
if "Field 5" in df.columns:
    def process_field5(value):
        match = re.match(r'^([\+\d\s\-]*)', value)
        if match:
            part = match.group(1)
            digits = re.sub(r'\D', '', part)
        else:
            digits = ""
        
        if len(digits) >= 4:
            last_four = digits[-4:]
            if last_four.isdigit() and 1950 <= int(last_four) <= 2025:
                digits = digits[:-4]
        
        if digits:
            digits = '+' + digits
        
        return digits

    df["Field 5"] = df["Field 5"].apply(process_field5)
    df['WhatsApp Link'] = df['Field 5'].apply(lambda x: f'https://wa.me/{x[1:]}' if x.startswith('+') else '')

# Step 9: Save to Excel
file_path = "restaurants.xlsx"
df.to_excel(file_path, index=False)

# Step 10: Format hyperlinks
wb = load_workbook(file_path)
ws = wb.active

for col_idx, cell in enumerate(ws[1], start=1):
    if cell.value == 'WhatsApp Link':
        link_col = col_idx
        break

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=link_col)
    url = cell.value
    if url:
        cell.hyperlink = url
        cell.value = "Open WhatsApp"
        cell.font = Font(color="0000FF", underline="single")

wb.save(file_path)

print(f"Excel file with clickable WhatsApp links has been saved as {file_path}")
