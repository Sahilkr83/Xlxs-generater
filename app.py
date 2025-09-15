import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font

def process_text(raw_text):
    # Split into lines and remove empty lines
    lines = raw_text.strip().split('\n')
    clean_lines = [line.strip() for line in lines if line.strip() != ""]

    # Remove lines containing 'Review'
    filtered_lines = [line for line in clean_lines if "Review" not in line]

    # Group lines where each group ends at 'Photos'
    grouped_data = []
    temp_group = []
    for line in filtered_lines:
        temp_group.append(line)
        if "Photos" in line:
            grouped_data.append(temp_group)
            temp_group = []
    if temp_group:
        grouped_data.append(temp_group)

    # Remove empty fields and shift
    processed_data = []
    for group in grouped_data:
        new_group = [field for field in group if field.strip() != ""]
        processed_data.append(new_group)

    # Find max columns after removing empty fields
    max_cols = max(len(group) for group in processed_data)

    # Pad with empty strings if necessary
    for group in processed_data:
        while len(group) < max_cols:
            group.append("")

    # Create DataFrame
    column_names = [f"Field {i+1}" for i in range(max_cols)]
    df = pd.DataFrame(processed_data, columns=column_names)

    # Remove 'Field 4' if it exists
    if "Field 4" in df.columns:
        df = df.drop(columns=["Field 4"])

    # Process 'Field 5'
    if "Field 5" in df.columns:
        def process_field5(value):
            match = re.match(r'^([\+\d\s\-]*)', value)
            if match:
                part = match.group(1)
                digits = re.sub(r'\D', '', part)
            else:
                digits = ""
            if len(digits) >= 4:
                last_four = digits[-4:]
                if last_four.isdigit() and 1950 <= int(last_four) <= 2025:
                    digits = digits[:-4]
            if digits:
                digits = '+' + digits
            return digits

        df["Field 5"] = df["Field 5"].apply(process_field5)
        df['WhatsApp Link'] = df["Field 5"].apply(lambda x: f'https://wa.me/{x[1:]}' if x.startswith('+') else '')

    # Remove empty columns
    df = df.dropna(axis=1, how='all')

    # Keep only first 5 columns and 'WhatsApp Link' if exists
    cols_to_keep = [col for col in df.columns[:5]]
    if 'WhatsApp Link' in df.columns:
        cols_to_keep.append('WhatsApp Link')
    df = df[cols_to_keep]

    # Rename columns based on expected info
    rename_map = {
        'Field 1': 'Name',
        'Field 2': 'Address',
        'Field 3': 'Description of Business',
        'Field 4': 'Phone Number',
        'Field 5': 'Phone Number'
    }
    df = df.rename(columns={col: rename_map.get(col, col) for col in df.columns})

    return df

# Streamlit UI
st.title("Restaurant Data Processor")

raw_text = st.text_area("Paste the raw text here", height=300)

if st.button("Process"):
    if raw_text.strip() == "":
        st.error("Please paste some text to process.")
    else:
        df = process_text(raw_text)
        st.success("Data processed!")
        st.dataframe(df)

        # Save Excel file to a BytesIO buffer
        output = BytesIO()
        df.to_excel(output, index=False)
        
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active

        # Find 'WhatsApp Link' column
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'WhatsApp Link':
                link_col = col_idx
                break
        else:
            link_col = None

        if link_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col)
                url = cell.value
                if url:
                    cell.hyperlink = url
                    cell.value = "Open WhatsApp"
                    cell.font = Font(color="0000FF", underline="single")

        # Save updated workbook to buffer
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # Download button
        st.download_button(
            label="Download Excel File",
            data=final_output,
            file_name="Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
